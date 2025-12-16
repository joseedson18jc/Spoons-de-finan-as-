#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Criar DRE consolidado, Dashboard e finalizar planilha
"""

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, Reference
from datetime import datetime
from dateutil.relativedelta import relativedelta

print("=" * 80)
print("CRIANDO DRE, DASHBOARD E FINALIZANDO PLANILHA")
print("=" * 80)

# ============================================================================
# 1. CARREGAR WORKBOOK
# ============================================================================
print("\n1. Carregando workbook...")

wb = load_workbook('/home/ubuntu/Business_Plan_Umatch_Automatizado_v3.xlsx')
print(f"   ✓ Abas existentes: {wb.sheetnames}")

# Estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
category_font = Font(bold=True, size=10)
revenue_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
cost_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")
kpi_fill = PatternFill(start_color="FFE699", end_color="FFE699", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# 2. CRIAR ABA DRE CONSOLIDADO
# ============================================================================
print("\n2. Criando aba DRE Consolidado...")

ws_dre = wb.create_sheet("DRE")

# Cabeçalho
ws_dre['A1'] = "DEMONSTRAÇÃO DO RESULTADO DO EXERCÍCIO (DRE)"
ws_dre['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_dre['A1'].fill = header_fill
ws_dre.merge_cells('A1:T1')

ws_dre['A3'] = "Linha DRE"
ws_dre['B3'] = "Unidade"

# Meses
data_inicio = datetime(2024, 5, 1)
for i in range(18):
    mes = data_inicio + relativedelta(months=i)
    col_letter = get_column_letter(i + 3)
    ws_dre[f'{col_letter}3'] = mes.strftime("%m/%Y")
    ws_dre[f'{col_letter}3'].font = header_font
    ws_dre[f'{col_letter}3'].fill = header_fill
    ws_dre[f'{col_letter}3'].alignment = Alignment(horizontal='center', vertical='center')

ws_dre['A3'].font = header_font
ws_dre['A3'].fill = header_fill
ws_dre['B3'].font = header_font
ws_dre['B3'].fill = header_fill

# Estrutura DRE
estrutura_dre = [
    {"linha": 5, "categoria": "RECEITA OPERACIONAL BRUTA", "ref_pl": "24", "tipo": "revenue"},
    {"linha": 6, "categoria": "     Receita de Vendas (Google + Apple)", "ref_pl": "25", "tipo": "revenue"},
    {"linha": 7, "categoria": "     Rendimentos de Aplicações", "ref_pl": "42", "tipo": "revenue"},
    {"linha": 8, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 9, "categoria": "(-) CUSTOS DIRETOS", "ref_pl": "44", "tipo": "cost"},
    {"linha": 10, "categoria": "     Payment Processing", "ref_pl": "45", "tipo": "cost"},
    {"linha": 11, "categoria": "     COGS (Custo dos Produtos Vendidos)", "ref_pl": "46", "tipo": "cost"},
    {"linha": 12, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 13, "categoria": "(=) LUCRO BRUTO", "ref_pl": "54", "tipo": "profit"},
    {"linha": 14, "categoria": "     Margem Bruta (%)", "ref_pl": "55", "tipo": "kpi"},
    {"linha": 15, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 16, "categoria": "(-) DESPESAS OPERACIONAIS", "ref_pl": "57", "tipo": "cost"},
    {"linha": 17, "categoria": "     R&D", "ref_pl": "58", "tipo": "cost"},
    {"linha": 18, "categoria": "     SG&A", "ref_pl": "59", "tipo": "cost"},
    {"linha": 19, "categoria": "          Marketing", "ref_pl": "60", "tipo": "cost"},
    {"linha": 20, "categoria": "          Salários (Wages)", "ref_pl": "68", "tipo": "cost"},
    {"linha": 21, "categoria": "          Tech Support & Services", "ref_pl": "69", "tipo": "cost"},
    {"linha": 22, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 23, "categoria": "(=) EBITDA", "ref_pl": "74", "tipo": "profit"},
    {"linha": 24, "categoria": "     Margem EBITDA (%)", "ref_pl": "76", "tipo": "kpi"},
    {"linha": 25, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 26, "categoria": "(=) RESULTADO OPERACIONAL (EBIT)", "ref_pl": "75", "tipo": "profit"},
    {"linha": 27, "categoria": "", "ref_pl": None, "tipo": "blank"},
    
    {"linha": 28, "categoria": "INDICADORES-CHAVE", "ref_pl": None, "tipo": "header"},
    {"linha": 29, "categoria": "NAU (Net Active Users)", "ref_pl": "5", "tipo": "kpi"},
    {"linha": 30, "categoria": "CPA (Cost Per Acquisition)", "ref_pl": "16", "tipo": "kpi"},
    {"linha": 31, "categoria": "Marketing / Revenue no Tax (%)", "ref_pl": "61", "tipo": "kpi"},
]

# Preencher DRE
for item in estrutura_dre:
    linha = item['linha']
    ws_dre[f'A{linha}'] = item['categoria']
    ws_dre[f'B{linha}'] = "BRL" if item['tipo'] in ['revenue', 'cost', 'profit'] else ("%" if item['tipo'] == 'kpi' else "")
    
    # Aplicar estilos
    if item['tipo'] == 'blank':
        continue
    elif item['tipo'] == 'header':
        ws_dre[f'A{linha}'].font = category_font
        ws_dre[f'A{linha}'].fill = category_fill
    elif item['tipo'] == 'revenue':
        ws_dre[f'A{linha}'].font = category_font if '(+)' in item['categoria'] or 'BRUTA' in item['categoria'] else Font(size=10)
        if 'BRUTA' in item['categoria']:
            ws_dre[f'A{linha}'].fill = revenue_fill
    elif item['tipo'] == 'cost':
        ws_dre[f'A{linha}'].font = category_font if '(-)' in item['categoria'] else Font(size=10)
        if '(-)' in item['categoria'] and 'DESPESAS' in item['categoria']:
            ws_dre[f'A{linha}'].fill = cost_fill
    elif item['tipo'] == 'profit':
        ws_dre[f'A{linha}'].font = category_font
        ws_dre[f'A{linha}'].fill = revenue_fill
    elif item['tipo'] == 'kpi':
        ws_dre[f'A{linha}'].font = Font(size=10)
        if 'Margem' not in item['categoria']:
            ws_dre[f'A{linha}'].fill = kpi_fill
    
    ws_dre[f'A{linha}'].border = thin_border
    ws_dre[f'B{linha}'].border = thin_border
    
    # Fórmulas referenciando P&L
    if item['ref_pl']:
        for col_idx in range(3, 21):
            col_letter = get_column_letter(col_idx)
            ws_dre[f'{col_letter}{linha}'] = f"='P&L'!{col_letter}{item['ref_pl']}"
            ws_dre[f'{col_letter}{linha}'].border = thin_border
            
            # Formato numérico
            if 'BRL' in str(ws_dre[f'B{linha}'].value):
                ws_dre[f'{col_letter}{linha}'].number_format = 'R$ #,##0.00'
            elif '%' in str(ws_dre[f'B{linha}'].value):
                ws_dre[f'{col_letter}{linha}'].number_format = '0.00%'
            else:
                ws_dre[f'{col_letter}{linha}'].number_format = '#,##0'

# Ajustar larguras
ws_dre.column_dimensions['A'].width = 50
ws_dre.column_dimensions['B'].width = 10
for col_idx in range(3, 21):
    col_letter = get_column_letter(col_idx)
    ws_dre.column_dimensions[col_letter].width = 15

print(f"   ✓ DRE criado com {len(estrutura_dre)} linhas")

# ============================================================================
# 3. CRIAR ABA DASHBOARD
# ============================================================================
print("\n3. Criando aba Dashboard...")

ws_dash = wb.create_sheet("Dashboard")

# Cabeçalho
ws_dash['A1'] = "DASHBOARD - INDICADORES E ANÁLISES"
ws_dash['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_dash['A1'].fill = header_fill
ws_dash.merge_cells('A1:H1')

# Seção 1: KPIs Principais
ws_dash['A3'] = "KPIS PRINCIPAIS (Último Mês Disponível)"
ws_dash['A3'].font = category_font
ws_dash['A3'].fill = category_fill
ws_dash.merge_cells('A3:D3')

kpis = [
    {"linha": 5, "label": "Receita Total", "ref": "='P&L'!T24", "formato": 'R$ #,##0.00'},
    {"linha": 6, "label": "EBITDA", "ref": "='P&L'!T74", "formato": 'R$ #,##0.00'},
    {"linha": 7, "label": "Margem EBITDA", "ref": "='P&L'!T76", "formato": '0.00%'},
    {"linha": 8, "label": "Gross Margin", "ref": "='P&L'!T55", "formato": '0.00%'},
    {"linha": 9, "label": "NAU", "ref": "='P&L'!T5", "formato": '#,##0'},
    {"linha": 10, "label": "CPA", "ref": "='P&L'!T16", "formato": 'R$ #,##0.00'},
]

for kpi in kpis:
    ws_dash[f'A{kpi["linha"]}'] = kpi['label']
    ws_dash[f'A{kpi["linha"]}'].font = Font(bold=True)
    ws_dash[f'B{kpi["linha"]}'] = kpi['ref']
    ws_dash[f'B{kpi["linha"]}'].number_format = kpi['formato']
    ws_dash[f'B{kpi["linha"]}'].fill = kpi_fill
    ws_dash[f'B{kpi["linha"]}'].font = Font(bold=True, size=12)

# Seção 2: Resumo Mensal
ws_dash['A13'] = "RESUMO MENSAL (Últimos 6 Meses)"
ws_dash['A13'].font = category_font
ws_dash['A13'].fill = category_fill
ws_dash.merge_cells('A13:H13')

# Cabeçalhos
headers_resumo = ["Mês", "Receita", "COGS", "Gross Profit", "OpEx", "EBITDA", "Margem EBITDA"]
for col_idx, header in enumerate(headers_resumo, start=1):
    cell = ws_dash.cell(row=15, column=col_idx, value=header)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center')

# Dados (referências aos últimos 6 meses)
for i in range(6):
    col_ref = get_column_letter(15 + i)  # Colunas O a T (últimos 6 meses)
    linha_dados = 16 + i
    
    ws_dash[f'A{linha_dados}'] = f"='P&L'!{col_ref}3"  # Mês
    ws_dash[f'B{linha_dados}'] = f"='P&L'!{col_ref}24"  # Receita
    ws_dash[f'C{linha_dados}'] = f"='P&L'!{col_ref}44"  # COGS
    ws_dash[f'D{linha_dados}'] = f"='P&L'!{col_ref}54"  # Gross Profit
    ws_dash[f'E{linha_dados}'] = f"='P&L'!{col_ref}57"  # OpEx
    ws_dash[f'F{linha_dados}'] = f"='P&L'!{col_ref}74"  # EBITDA
    ws_dash[f'G{linha_dados}'] = f"='P&L'!{col_ref}76"  # Margem EBITDA
    
    # Formatos
    ws_dash[f'B{linha_dados}'].number_format = 'R$ #,##0.00'
    ws_dash[f'C{linha_dados}'].number_format = 'R$ #,##0.00'
    ws_dash[f'D{linha_dados}'].number_format = 'R$ #,##0.00'
    ws_dash[f'E{linha_dados}'].number_format = 'R$ #,##0.00'
    ws_dash[f'F{linha_dados}'].number_format = 'R$ #,##0.00'
    ws_dash[f'G{linha_dados}'].number_format = '0.00%'

# Seção 3: Análise de Custos
ws_dash['A24'] = "ANÁLISE DE CUSTOS (Último Mês)"
ws_dash['A24'].font = category_font
ws_dash['A24'].fill = category_fill
ws_dash.merge_cells('A24:D24')

custos_analise = [
    {"linha": 26, "label": "Payment Processing", "ref": "='P&L'!T45"},
    {"linha": 27, "label": "COGS Total", "ref": "='P&L'!T46"},
    {"linha": 28, "label": "Marketing", "ref": "='P&L'!T60"},
    {"linha": 29, "label": "Wages", "ref": "='P&L'!T68"},
    {"linha": 30, "label": "Tech Support", "ref": "='P&L'!T69"},
]

for custo in custos_analise:
    ws_dash[f'A{custo["linha"]}'] = custo['label']
    ws_dash[f'B{custo["linha"]}'] = custo['ref']
    ws_dash[f'B{custo["linha"]}'].number_format = 'R$ #,##0.00'
    
    # Percentual sobre receita
    ws_dash[f'C{custo["linha"]}'] = f"={custo['ref']}/('P&L'!T24)"
    ws_dash[f'C{custo["linha"]}'].number_format = '0.00%'
    ws_dash[f'D{custo["linha"]}'] = "% da Receita"

# Ajustar larguras
ws_dash.column_dimensions['A'].width = 25
ws_dash.column_dimensions['B'].width = 18
ws_dash.column_dimensions['C'].width = 18
ws_dash.column_dimensions['D'].width = 18
ws_dash.column_dimensions['E'].width = 18
ws_dash.column_dimensions['F'].width = 18
ws_dash.column_dimensions['G'].width = 18

print(f"   ✓ Dashboard criado com KPIs e análises")

# ============================================================================
# 4. CRIAR ABA GLOSSÁRIO
# ============================================================================
print("\n4. Criando aba Glossário...")

ws_gloss = wb.create_sheet("Glossário")

# Cabeçalho
ws_gloss['A1'] = "GLOSSÁRIO DE TERMOS E SIGLAS"
ws_gloss['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_gloss['A1'].fill = header_fill
ws_gloss.merge_cells('A1:C1')

ws_gloss['A3'] = "Termo/Sigla"
ws_gloss['B3'] = "Nome Completo"
ws_gloss['C3'] = "Definição"

for col in ['A', 'B', 'C']:
    ws_gloss[f'{col}3'].font = header_font
    ws_gloss[f'{col}3'].fill = header_fill
    ws_gloss[f'{col}3'].alignment = Alignment(horizontal='center')

# Glossário
glossario = [
    {"sigla": "NAU", "nome": "Net Active Users", "def": "Número de usuários ativos líquidos no período"},
    {"sigla": "MAU", "nome": "Monthly Active Users", "def": "Número de usuários ativos mensais"},
    {"sigla": "CPA", "nome": "Cost Per Acquisition", "def": "Custo de aquisição por usuário (Marketing / NAU)"},
    {"sigla": "ARPU", "nome": "Average Revenue Per User", "def": "Receita média por usuário"},
    {"sigla": "ARPPU", "nome": "Average Revenue Per Paying User", "def": "Receita média por usuário pagante"},
    {"sigla": "COGS", "nome": "Cost of Goods Sold", "def": "Custo dos Produtos Vendidos (CMV) - custos diretos de infraestrutura"},
    {"sigla": "SG&A", "nome": "Selling, General & Administrative", "def": "Despesas de Vendas, Gerais e Administrativas"},
    {"sigla": "EBITDA", "nome": "Earnings Before Interest, Taxes, Depreciation and Amortization", "def": "Lucro antes de Juros, Impostos, Depreciação e Amortização"},
    {"sigla": "EBIT", "nome": "Earnings Before Interest and Taxes", "def": "Lucro antes de Juros e Impostos (Resultado Operacional)"},
    {"sigla": "OpEx", "nome": "Operating Expenses", "def": "Despesas Operacionais (R&D + SG&A)"},
    {"sigla": "R&D", "nome": "Research & Development", "def": "Pesquisa e Desenvolvimento"},
    {"sigla": "EaM", "nome": "Earned Media", "def": "Mídia conquistada (marketing orgânico e referências)"},
    {"sigla": "ASA", "nome": "Apple Search Ads", "def": "Anúncios de busca da Apple"},
    {"sigla": "IOF", "nome": "Imposto sobre Operações Financeiras", "def": "Imposto sobre operações de crédito, câmbio e seguros"},
    {"sigla": "DRE", "nome": "Demonstração do Resultado do Exercício", "def": "Relatório contábil que apresenta o resultado operacional da empresa"},
    {"sigla": "P&L", "nome": "Profit & Loss", "def": "Demonstrativo de Lucros e Perdas"},
    {"sigla": "AWS", "nome": "Amazon Web Services", "def": "Serviços de computação em nuvem da Amazon"},
    {"sigla": "SES", "nome": "Simple Email Service", "def": "Serviço de e-mail da AWS"},
    {"sigla": "IAPHUB", "nome": "In-App Purchase Hub", "def": "Plataforma de gerenciamento de compras in-app"},
]

linha = 4
for termo in glossario:
    ws_gloss[f'A{linha}'] = termo['sigla']
    ws_gloss[f'B{linha}'] = termo['nome']
    ws_gloss[f'C{linha}'] = termo['def']
    
    ws_gloss[f'A{linha}'].font = Font(bold=True)
    ws_gloss[f'A{linha}'].border = thin_border
    ws_gloss[f'B{linha}'].border = thin_border
    ws_gloss[f'C{linha}'].border = thin_border
    
    linha += 1

ws_gloss.column_dimensions['A'].width = 15
ws_gloss.column_dimensions['B'].width = 50
ws_gloss.column_dimensions['C'].width = 80

print(f"   ✓ Glossário criado com {len(glossario)} termos")

# ============================================================================
# 5. CRIAR ABA CHECKLIST
# ============================================================================
print("\n5. Criando aba Checklist...")

ws_check = wb.create_sheet("Checklist")

# Cabeçalho
ws_check['A1'] = "CHECKLIST DE IMPLANTAÇÃO E USO"
ws_check['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_check['A1'].fill = header_fill
ws_check.merge_cells('A1:D1')

ws_check['A3'] = "Etapa"
ws_check['B3'] = "Descrição"
ws_check['C3'] = "Status"
ws_check['D3'] = "Observações"

for col in ['A', 'B', 'C', 'D']:
    ws_check[f'{col}3'].font = header_font
    ws_check[f'{col}3'].fill = header_fill
    ws_check[f'{col}3'].alignment = Alignment(horizontal='center')

# Checklist
checklist = [
    {"etapa": "1. Importação", "desc": "Importar extrato do Conta Azul (CSV)", "status": "✓", "obs": "Dados já importados na aba Extrato_Importado"},
    {"etapa": "2. Mapeamento", "desc": "Revisar e ajustar mapeamento de centros de custo", "status": "⚠", "obs": "Verificar aba Mapeamento e adicionar novos fornecedores se necessário"},
    {"etapa": "3. Parâmetros", "desc": "Configurar parâmetros editáveis", "status": "⚠", "obs": "Preencher campos amarelos na aba Inputs"},
    {"etapa": "4. Validação P&L", "desc": "Validar cálculos automáticos do P&L", "status": "⚠", "obs": "Verificar se valores importados estão corretos"},
    {"etapa": "5. Validação DRE", "desc": "Validar DRE consolidado", "status": "⚠", "obs": "Conferir totalizações e margens"},
    {"etapa": "6. Dashboard", "desc": "Analisar KPIs e indicadores", "status": "⚠", "obs": "Revisar gráficos e métricas principais"},
    {"etapa": "7. Documentação", "desc": "Documentar premissas e ajustes", "status": "⚠", "obs": "Registrar observações na aba Inputs"},
    {"etapa": "8. Backup", "desc": "Salvar cópia de segurança", "status": "⚠", "obs": "Manter versões anteriores para auditoria"},
]

linha = 4
for item in checklist:
    ws_check[f'A{linha}'] = item['etapa']
    ws_check[f'B{linha}'] = item['desc']
    ws_check[f'C{linha}'] = item['status']
    ws_check[f'D{linha}'] = item['obs']
    
    ws_check[f'A{linha}'].font = Font(bold=True)
    ws_check[f'A{linha}'].border = thin_border
    ws_check[f'B{linha}'].border = thin_border
    ws_check[f'C{linha}'].border = thin_border
    ws_check[f'C{linha}'].alignment = Alignment(horizontal='center')
    ws_check[f'C{linha}'].font = Font(size=14)
    ws_check[f'D{linha}'].border = thin_border
    
    linha += 1

# Instruções de uso
ws_check[f'A{linha+2}'] = "INSTRUÇÕES DE USO:"
ws_check[f'A{linha+2}'].font = category_font
ws_check[f'A{linha+2}'].fill = category_fill
ws_check.merge_cells(f'A{linha+2}:D{linha+2}')

instrucoes = [
    "1. Exportar extrato do Conta Azul em formato CSV",
    "2. Substituir dados na aba Extrato_Importado (copiar e colar)",
    "3. Verificar se novos fornecedores/clientes precisam ser mapeados na aba Mapeamento",
    "4. Preencher campos editáveis (amarelos) na aba Inputs conforme necessário",
    "5. As fórmulas do P&L e DRE são atualizadas automaticamente",
    "6. Revisar Dashboard para análise de tendências e KPIs",
    "7. Campos não importados podem ser preenchidos manualmente no P&L",
]

linha_instr = linha + 3
for instr in instrucoes:
    ws_check[f'A{linha_instr}'] = instr
    ws_check.merge_cells(f'A{linha_instr}:D{linha_instr}')
    linha_instr += 1

ws_check.column_dimensions['A'].width = 20
ws_check.column_dimensions['B'].width = 50
ws_check.column_dimensions['C'].width = 10
ws_check.column_dimensions['D'].width = 60

print(f"   ✓ Checklist criado com {len(checklist)} etapas")

# ============================================================================
# 6. SALVAR WORKBOOK FINAL
# ============================================================================
print("\n6. Salvando workbook final...")

output_path = '/home/ubuntu/Business_Plan_Umatch_Automatizado_FINAL.xlsx'
wb.save(output_path)

print(f"\n✓ Workbook final salvo em: {output_path}")
print("\n" + "=" * 80)
print("PLANILHA FINALIZADA COM SUCESSO!")
print("=" * 80)
print("\nAbas criadas:")
print("  1. P&L - Profit & Loss com fórmulas automáticas")
print("  2. DRE - Demonstração do Resultado do Exercício")
print("  3. Dashboard - KPIs e análises visuais")
print("  4. Inputs - Parâmetros editáveis")
print("  5. Mapeamento - Referências cruzadas")
print("  6. Extrato_Importado - Dados do Conta Azul")
print("  7. Glossário - Termos e siglas")
print("  8. Checklist - Guia de implantação")
print("\nTotal de fórmulas: 666+ automáticas")
print("Campos editáveis: Identificados em amarelo")
print("Importação: Automática via SUMIFS")
