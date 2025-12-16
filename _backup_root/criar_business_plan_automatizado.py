#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Criação do Business Plan Umatch Automatizado
Integração com Conta Azul em regime de competência
"""

import pandas as pd
import numpy as np
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.datavalidation import DataValidation
import json

print("=" * 80)
print("CRIAÇÃO DO BUSINESS PLAN UMATCH AUTOMATIZADO")
print("=" * 80)

# ============================================================================
# 1. CARREGAR DADOS ORIGINAIS
# ============================================================================
print("\n1. Carregando dados originais...")

# Carregar P&L original
pl_original = pd.read_csv('/home/ubuntu/upload/00_Business_Plan_Umatch.xlsx-P&L.csv')
print(f"   ✓ P&L: {pl_original.shape}")

# Carregar Assumptions
assumptions = pd.read_csv('/home/ubuntu/upload/00_Business_Plan_Umatch.xlsx-Assumptions.csv')
print(f"   ✓ Assumptions: {assumptions.shape}")

# Carregar extrato Conta Azul
extrato_raw = pd.read_csv('/home/ubuntu/upload/Extratodemovimentações-2025-ExtratoFinanceiro.csv')
print(f"   ✓ Extrato Conta Azul: {extrato_raw.shape}")

# ============================================================================
# 2. PROCESSAR EXTRATO CONTA AZUL
# ============================================================================
print("\n2. Processando extrato Conta Azul...")

# Converter data de competência
extrato_raw['Data de competência'] = pd.to_datetime(
    extrato_raw['Data de competência'], 
    format='%d/%m/%Y', 
    errors='coerce'
)

# Converter valores para numérico
def converter_valor_br(valor_str):
    """Converte valor brasileiro (R$ 1.234,56) para float"""
    if pd.isna(valor_str) or valor_str == '':
        return 0.0
    valor_str = str(valor_str).replace('R$', '').replace('.', '').replace(',', '.').strip()
    try:
        return float(valor_str)
    except:
        return 0.0

extrato_raw['Valor_Num'] = extrato_raw['Valor (R$)'].apply(converter_valor_br)
extrato_raw['Mes_Competencia'] = extrato_raw['Data de competência'].dt.to_period('M')

print(f"   ✓ Período: {extrato_raw['Data de competência'].min()} a {extrato_raw['Data de competência'].max()}")
print(f"   ✓ Total de movimentações: {len(extrato_raw)}")

# ============================================================================
# 3. CRIAR WORKBOOK EXCEL
# ============================================================================
print("\n3. Criando estrutura do workbook...")

wb = Workbook()

# Remover sheet padrão
if 'Sheet' in wb.sheetnames:
    wb.remove(wb['Sheet'])

# Definir estilos
header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
header_font = Font(bold=True, color="FFFFFF", size=11)
subheader_fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
subheader_font = Font(bold=True, size=10)
category_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
category_font = Font(bold=True, size=10)
editable_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
formula_fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
revenue_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")
cost_fill = PatternFill(start_color="F8CBAD", end_color="F8CBAD", fill_type="solid")

thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# ============================================================================
# 4. ABA 1: INPUTS (Dados Editáveis e Importados)
# ============================================================================
print("\n4. Criando aba INPUTS...")

ws_inputs = wb.create_sheet("Inputs")

# Cabeçalho
ws_inputs['A1'] = "DADOS DE ENTRADA - BUSINESS PLAN UMATCH"
ws_inputs['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_inputs['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_inputs.merge_cells('A1:E1')

ws_inputs['A3'] = "Seção"
ws_inputs['B3'] = "Parâmetro"
ws_inputs['C3'] = "Valor"
ws_inputs['D3'] = "Unidade"
ws_inputs['E3'] = "Observações"

for col in ['A', 'B', 'C', 'D', 'E']:
    ws_inputs[f'{col}3'].font = header_font
    ws_inputs[f'{col}3'].fill = header_fill
    ws_inputs[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')

# Parâmetros editáveis
parametros = [
    ["Receita", "Taxa Apple/Google", "0.85", "%", "Receita líquida após desconto de 15%"],
    ["Receita", "Receita Bruta = Líquida / 0.85", "=C4/C4", "Fórmula", "Cálculo automático"],
    ["", "", "", "", ""],
    ["Custos", "COGS - AWS", "", "R$", "Custo AWS (editável)"],
    ["Custos", "COGS - Cloudflare", "", "R$", "Custo Cloudflare (editável)"],
    ["Custos", "COGS - Heroku", "", "R$", "Custo Heroku (editável)"],
    ["Custos", "COGS - IAPHUB", "", "R$", "Custo IAPHUB (editável)"],
    ["Custos", "COGS - MailGun", "", "R$", "Custo MailGun (editável)"],
    ["Custos", "COGS - AWS SES", "", "R$", "Custo AWS SES (editável)"],
    ["", "", "", "", ""],
    ["SG&A", "Marketing", "", "R$", "Despesas de Marketing (editável)"],
    ["SG&A", "Wages", "", "R$", "Salários e Pró-labore (editável)"],
    ["SG&A", "Tech Support & Services", "", "R$", "Serviços de tecnologia (editável)"],
    ["", "", "", "", ""],
    ["Importação", "Status Importação", "Pendente", "Status", "Atualizado após importação"],
    ["Importação", "Última Atualização", "", "Data", "Data da última importação"],
    ["Importação", "Total Registros", "", "Qtd", "Quantidade de registros importados"],
]

row = 4
for param in parametros:
    for col_idx, val in enumerate(param, start=1):
        cell = ws_inputs.cell(row=row, column=col_idx, value=val)
        if col_idx == 3 and val == "":  # Células editáveis
            cell.fill = editable_fill
        cell.border = thin_border
    row += 1

# Ajustar larguras
ws_inputs.column_dimensions['A'].width = 15
ws_inputs.column_dimensions['B'].width = 35
ws_inputs.column_dimensions['C'].width = 20
ws_inputs.column_dimensions['D'].width = 12
ws_inputs.column_dimensions['E'].width = 40

# ============================================================================
# 5. ABA 2: MAPEAMENTO (Referência Cruzada)
# ============================================================================
print("\n5. Criando aba MAPEAMENTO...")

ws_map = wb.create_sheet("Mapeamento")

# Cabeçalho
ws_map['A1'] = "MAPEAMENTO DE CENTROS DE CUSTO E FORNECEDORES"
ws_map['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_map['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_map.merge_cells('A1:G1')

ws_map['A3'] = "Grupo Financeiro"
ws_map['B3'] = "Centro de Custo (Conta Azul)"
ws_map['C3'] = "Fornecedor/Cliente"
ws_map['D3'] = "Linha P&L"
ws_map['E3'] = "Tipo"
ws_map['F3'] = "Ativo"
ws_map['G3'] = "Observações"

for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
    ws_map[f'{col}3'].font = header_font
    ws_map[f'{col}3'].fill = header_fill
    ws_map[f'{col}3'].alignment = Alignment(horizontal='center', vertical='center')

# Mapeamento de grupos financeiros
mapeamentos = [
    # RECEITAS
    ["Receita Google", "Google Play Net Revenue", "GOOGLE BRASIL PAGAMENTOS LTDA", "25", "Receita", "Sim", "Receita Google Play"],
    ["Receita Apple", "App Store Net Revenue", "App Store (Apple)", "33", "Receita", "Sim", "Receita App Store"],
    ["Receita Brasil", "Google Play Net Revenue", "GOOGLE BRASIL PAGAMENTOS LTDA", "26", "Receita", "Sim", "Receita Brasil - Google"],
    ["Receita Brasil", "App Store Net Revenue", "App Store (Apple)", "34", "Receita", "Sim", "Receita Brasil - Apple"],
    ["Receita USA", "Google Play Net Revenue", "GOOGLE BRASIL PAGAMENTOS LTDA", "28", "Receita", "Sim", "Receita USA - Google"],
    ["Receita USA", "App Store Net Revenue", "App Store (Apple)", "36", "Receita", "Sim", "Receita USA - Apple"],
    ["", "", "", "", "", "", ""],
    
    # COGS
    ["COGS", "Web Services Expenses", "AWS", "43", "Custo", "Sim", "Amazon Web Services"],
    ["COGS", "Web Services Expenses", "Cloudflare", "44", "Custo", "Sim", "Cloudflare"],
    ["COGS", "Web Services Expenses", "Heroku", "45", "Custo", "Sim", "Heroku"],
    ["COGS", "Web Services Expenses", "IAPHUB", "46", "Custo", "Sim", "IAPHUB"],
    ["COGS", "Web Services Expenses", "MailGun", "47", "Custo", "Sim", "MailGun"],
    ["COGS", "Web Services Expenses", "AWS SES", "48", "Custo", "Sim", "AWS SES"],
    ["", "", "", "", "", "", ""],
    
    # SG&A
    ["SG&A", "Marketing & Growth Expenses", "MGA MARKETING LTDA", "56", "Despesa", "Sim", "Marketing"],
    ["SG&A", "Marketing & Growth Expenses", "Diversos", "56", "Despesa", "Sim", "Marketing - Diversos"],
    ["SG&A", "Wages Expenses", "Diversos", "64", "Despesa", "Sim", "Salários e Pró-labore"],
    ["SG&A", "Tech Support & Services", "Adobe", "68", "Despesa", "Sim", "Adobe Creative Cloud"],
    ["SG&A", "Tech Support & Services", "Canva", "68", "Despesa", "Sim", "Canva"],
    ["SG&A", "Tech Support & Services", "ClickSign", "68", "Despesa", "Sim", "ClickSign"],
    ["SG&A", "Tech Support & Services", "COMPANYHERO SAO PAULO BRA", "68", "Despesa", "Sim", "CompanyHero"],
    ["SG&A", "Tech Support & Services", "Diversos", "65", "Despesa", "Sim", "Tech Support - Diversos"],
    ["", "", "", "", "", "", ""],
    
    # OUTRAS DESPESAS
    ["Outras Despesas", "Legal & Accounting Expenses", "BHUB.AI", "90", "Despesa", "Sim", "BPO Financeiro"],
    ["Outras Despesas", "Legal & Accounting Expenses", "WOLFF E SCRIPES ADVOGADOS", "90", "Despesa", "Sim", "Honorários Advocatícios"],
    ["Outras Despesas", "Office Expenses", "GO OFFICES LATAM S/A", "90", "Despesa", "Sim", "Aluguel"],
    ["Outras Despesas", "Office Expenses", "CO-SERVICES DO BRASIL  SERVICOS COMBINADOS DE APOIO A EDIFICIOS LTDA", "90", "Despesa", "Sim", "Serviços de Escritório"],
    ["Outras Despesas", "Travel", "American Airlines", "90", "Despesa", "Sim", "Viagens"],
    ["Outras Despesas", "Other Taxes", "IMPOSTOS/TRIBUTOS", "90", "Despesa", "Sim", "Impostos e Tributos"],
    ["Outras Despesas", "Payroll Tax - Brazil", "IMPOSTOS/TRIBUTOS", "90", "Despesa", "Sim", "Impostos sobre Folha"],
    ["", "", "", "", "", "", ""],
    
    # RENDIMENTOS
    ["Rendimentos", "Rendimentos de Aplicações", "CONTA SIMPLES", "38", "Receita", "Sim", "Rendimentos CDI - Conta Simples"],
    ["Rendimentos", "Rendimentos de Aplicações", "BANCO INTER", "38", "Receita", "Sim", "Rendimentos - Banco Inter"],
]

row = 4
for mapa in mapeamentos:
    for col_idx, val in enumerate(mapa, start=1):
        cell = ws_map.cell(row=row, column=col_idx, value=val)
        cell.border = thin_border
        if col_idx in [1, 2, 3]:  # Colunas editáveis
            cell.fill = editable_fill
    row += 1

# Ajustar larguras
ws_map.column_dimensions['A'].width = 20
ws_map.column_dimensions['B'].width = 35
ws_map.column_dimensions['C'].width = 40
ws_map.column_dimensions['D'].width = 12
ws_map.column_dimensions['E'].width = 12
ws_map.column_dimensions['F'].width = 10
ws_map.column_dimensions['G'].width = 35

print(f"   ✓ {len(mapeamentos)} mapeamentos criados")

# ============================================================================
# 6. ABA 3: EXTRATO IMPORTADO
# ============================================================================
print("\n6. Criando aba EXTRATO IMPORTADO...")

ws_extrato = wb.create_sheet("Extrato_Importado")

# Cabeçalho
ws_extrato['A1'] = "EXTRATO CONTA AZUL - DADOS IMPORTADOS"
ws_extrato['A1'].font = Font(bold=True, size=14, color="FFFFFF")
ws_extrato['A1'].fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
ws_extrato.merge_cells('A1:J1')

# Colunas principais
colunas_extrato = [
    "Data Competência",
    "Centro de Custo",
    "Fornecedor/Cliente",
    "Descrição",
    "Valor (R$)",
    "Tipo Operação",
    "Categoria",
    "Mês",
    "Grupo Mapeado",
    "Linha P&L"
]

for col_idx, col_name in enumerate(colunas_extrato, start=1):
    cell = ws_extrato.cell(row=3, column=col_idx, value=col_name)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = Alignment(horizontal='center', vertical='center')
    cell.border = thin_border

# Inserir dados do extrato
row = 4
for idx, registro in extrato_raw.iterrows():
    ws_extrato.cell(row=row, column=1, value=registro['Data de competência'])
    ws_extrato.cell(row=row, column=2, value=registro['Centro de Custo 1'])
    ws_extrato.cell(row=row, column=3, value=registro['Nome do fornecedor/cliente'])
    ws_extrato.cell(row=row, column=4, value=registro['Descrição'])
    ws_extrato.cell(row=row, column=5, value=registro['Valor_Num'])
    ws_extrato.cell(row=row, column=6, value=registro['Tipo da operação'])
    ws_extrato.cell(row=row, column=7, value=registro['Categoria 1'])
    ws_extrato.cell(row=row, column=8, value=str(registro['Mes_Competencia']))
    
    # Aplicar bordas
    for col in range(1, 11):
        ws_extrato.cell(row=row, column=col).border = thin_border
    
    row += 1
    if row > 1000:  # Limitar para não sobrecarregar
        break

# Ajustar larguras
ws_extrato.column_dimensions['A'].width = 18
ws_extrato.column_dimensions['B'].width = 35
ws_extrato.column_dimensions['C'].width = 45
ws_extrato.column_dimensions['D'].width = 50
ws_extrato.column_dimensions['E'].width = 15
ws_extrato.column_dimensions['F'].width = 15
ws_extrato.column_dimensions['G'].width = 35
ws_extrato.column_dimensions['H'].width = 12
ws_extrato.column_dimensions['I'].width = 20
ws_extrato.column_dimensions['J'].width = 12

print(f"   ✓ {min(len(extrato_raw), 997)} registros importados")

# ============================================================================
# 7. SALVAR WORKBOOK
# ============================================================================
print("\n7. Salvando workbook...")

output_path = '/home/ubuntu/Business_Plan_Umatch_Automatizado_v1.xlsx'
wb.save(output_path)

print(f"\n✓ Workbook salvo em: {output_path}")
print("\n" + "=" * 80)
print("FASE 1 CONCLUÍDA - Estrutura base criada")
print("=" * 80)
print("\nPróximas etapas:")
print("  - Criar aba P&L com fórmulas automáticas")
print("  - Criar aba DRE consolidado")
print("  - Criar aba Dashboard com gráficos")
print("  - Criar aba Glossário")
print("  - Criar aba Checklist")
